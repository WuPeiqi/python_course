from openpyxl import load_workbook

wb = load_workbook("files/p1.xlsx")

# sheet相关操作

# 1.获取excel文件中的所有sheet名称
"""
print(wb.sheetnames) # ['数据导出', '用户列表', 'Sheet1', 'Sheet2']
"""

# 2.选择sheet，基于sheet名称
"""
sheet = wb["数据导出"]
cell = sheet.cell(1, 2)
print(cell.value)
"""

# 3.选择sheet，基于索引位置
"""
sheet = wb.worksheets[0]
cell = sheet.cell(1,2)
print(cell.value)
"""

# 4.循环所有的sheet
"""
for name in wb.sheetnames:
    sheet = wb[name]
    cell = sheet.cell(1, 1)
    print(cell.value)
"""
"""
for sheet in wb.worksheets:
    cell = sheet.cell(1, 1)
    print(cell.value)
"""
"""
for sheet in wb:
    cell = sheet.cell(1, 1)
    print(cell.value)
"""