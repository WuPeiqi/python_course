"""
day09 作业题讲解
"""

# 1.for循环
"""
data_list = [11,22,33,44,55]
for item in data_list:
    print(item)
    break
else:
    print("else中的内容") # for循环中的内容全部执行了一遍，且未遇到break
"""

# 2.enumerate
"""
data_list = [11, 22, 33, 44, 55]
for index in range(len(data_list)):
    print(index+1, data_list[index])
"""
"""
data_list = [11, 22, 33, 44, 55]
for i, item in enumerate(data_list, 1):
    print(i, item)
"""

# ############### 1. 基于csv的用户注册和认证 #############
"""
import os

# 文件路径处理
base_dir = os.path.dirname(os.path.abspath(__file__))
db_file_path = os.path.join(base_dir, 'db.csv')

# 用户注册
while True:
    choice = input("是否进行用户注册（Y/N）？")
    choice = choice.upper()
    if choice not in {'Y', 'N'}:
        print('输入格式错误，请重新输入。')
        continue

    if choice == "N":
        break

    with open(db_file_path, mode='a', encoding='utf-8') as file_object:
        while True:
            user = input("请输入用户名（Q/q退出）：")
            if user.upper() == 'Q':
                break
            pwd = input("请输入密码：")
            file_object.write("{},{}\n".format(user, pwd))
            file_object.flush()
    break

# 用户登录
print("欢迎使用xx系统，请登录！")
username = input("请输入用户名：")
password = input("请输入密码：")
if not os.path.exists(db_file_path):
    print("用户文件不存在")
else:
    with open(db_file_path, mode='r', encoding='utf-8') as file_object:
        for line in file_object:
            user, pwd = line.strip().split(',')
            if username == user and pwd == password:
                print('登录成功')
                break
        else:
            print("用户名或密码错误")
"""

# ############### 2. 实现去网上获取指定地区的天气信息，并写入到Excel中。 #############
"""
import os
import requests
from xml.etree import ElementTree as ET
from openpyxl import workbook

# 处理文件路径
base_dir = os.path.dirname(os.path.abspath(__file__))
target_excel_file_path = os.path.join(base_dir, 'weather.xlsx')

# 创建excel且默认会创建一个sheet（名称为Sheet）
wb = workbook.Workbook()
del wb['Sheet']

while True:
    # 用户输入城市，并获取该城市的天气信息
    city = input("请输入城市（Q/q退出）：")
    if city.upper() == "Q":
        break
    url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
    res = requests.get(url=url)

    # 1.提取XML格式中的数据
    root = ET.XML(res.text)

    # 2.为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。
    sheet = wb.create_sheet(city)

    for row_index, node in enumerate(root, 1):
        text = node.text
        cell = sheet.cell(row_index, 1)
        cell.value = text

wb.save(target_excel_file_path)
"""

# ############### 3. 读取ini文件，并写入到Excel #############
"""
import os
import configparser
from openpyxl import workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# 文件路径处理
base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir, 'files', 'my.ini')
target_excel_file_path = os.path.join(base_dir, 'my.xlsx')

# 创建excel且默认会创建一个sheet（名称为Sheet）
wb = workbook.Workbook()
del wb['Sheet']

# 解析ini格式文件
config = configparser.ConfigParser()
config.read(file_path, encoding='utf-8')

# 循环获取每个节点，并为每个节点创建一个sheet
for section in config.sections():
    # 在excel中创建一个sheet，名称为ini文件的节点名称
    sheet = wb.create_sheet(section)

    # 边框和居中（表头和内容都需要）
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)

    align = Alignment(horizontal='center', vertical='center')

    # 为此在sheet设置表头
    title_dict = {"A1": "键", "B1": "值"}
    for position, text in title_dict.items():
        cell = sheet[position]
        # 设置值
        cell.value = text
        # 设置居中
        cell.alignment = align
        # 设置背景色
        cell.fill = PatternFill("solid", fgColor="6495ED")
        # 设置字体颜色
        cell.font = Font(name="微软雅黑", color="FFFFFF")
        # 设置边框
        cell.border = border

    # 读取此节点下的所有键值，并将键值写入到当前sheet中
    # row_index = 2
    # for key, val in config.items(section):
    #     c1 = sheet.cell(row_index, 1)
    #     c1.value = key
    #     c1.alignment = align
    #     c1.border = border
    #
    #     c2 = sheet.cell(row_index, 2)
    #     c2.value = val
    #     c2.alignment = align
    #     c2.border = border
    #     row_index += 1

    row_index = 2
    for group in config.items(section):
        # group = ("datadir","/var/lib/mysql")
        for col, text in enumerate(group, 1):
            cell = sheet.cell(row_index, col)
            cell.alignment = align
            cell.border = border
            cell.value = text
        row_index += 1

wb.save(target_excel_file_path)
"""

# ############### 4. 下载zip文件，并解压到指定路径 #############
import os
import shutil
import requests

# 文件路径处理
base_dir = os.path.dirname(os.path.abspath(__file__))
download_folder = os.path.join(base_dir, 'files', 'package')
if not os.path.exists(download_folder):
    os.makedirs(download_folder)

# 1.下载文件
file_url = 'https://files.cnblogs.com/files/wupeiqi/HtmlStore.zip'
res = requests.get(url=file_url)

# 2.将下载的文件保存到当前执行脚本同级目录下 /files/package/ 目录下（且文件名为 HtmlStore.zip ）
file_name = file_url.split('/')[-1]
zip_file_path = os.path.join(download_folder, file_name) # .../files/package/HtmlStore.zip
with open(zip_file_path, mode='wb') as file_object:
    file_object.write(res.content)

# 3.在将下载下来的文件解压到 /files/html/ 目录下
unpack_folder = os.path.join(base_dir, 'files', 'html')
shutil.unpack_archive(filename=zip_file_path, extract_dir=unpack_folder, format='zip')

