from openpyxl import load_workbook

wb = load_workbook("files/p1.xlsx")
sheet = wb.worksheets[0]

# 1.获取第N行第N列的单元格(位置是从1开始）
"""
cell = sheet.cell(1, 1)

print(cell.value)
print(cell.style)
print(cell.font)
print(cell.alignment)
"""

# 2.获取某个单元格
"""
c1 = sheet["A2"]
print(c1.value)

c2 = sheet['D4']
print(c2.value)
"""

# 3.第N行所有的单元格
"""
for cell in sheet[1]:
    print(cell.value)
"""

# 4.所有行的数据（获取某一列数据）
"""
for row in sheet.rows:
    print(row[0].value, row[1].value)
"""

# 5.获取所有列的数据
"""
for col in sheet.columns:
    print(col[1].value)
"""