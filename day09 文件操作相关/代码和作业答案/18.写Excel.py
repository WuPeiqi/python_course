from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, GradientFill


wb = load_workbook('files/p1.xlsx')
sheet = wb.worksheets[0]

sheet.print_area = "A1:D200"
sheet.print_title_cols = "A:D"
sheet.print_title_rows = "1:1"

wb.save("files/p2.xlsx")