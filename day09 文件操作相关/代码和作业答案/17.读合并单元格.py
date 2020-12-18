from openpyxl import load_workbook

wb = load_workbook("files/p1.xlsx")
sheet = wb.worksheets[2]

# 获取第N行第N列的单元格(位置是从1开始）
c1 = sheet.cell(1, 1)
print(c1)  # <MergedCell 'Sheet1'.B1>

c2 = sheet.cell(1, 2)
print(c2)  # <Cell 'Sheet1'.A1>
