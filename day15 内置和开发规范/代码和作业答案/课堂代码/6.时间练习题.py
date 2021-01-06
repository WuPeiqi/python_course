import os
import hashlib
from datetime import datetime

from openpyxl import load_workbook, workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = "db.xlsx"


def md5(origin):
    hash_object = hashlib.md5("sdfsdfsdfsd23sd".encode('utf-8'))
    hash_object.update(origin.encode('utf-8'))
    return hash_object.hexdigest()


def register(username, password):
    db_file_path = os.path.join(BASE_DIR, FILE_NAME)
    if os.path.exists(db_file_path):
        wb = load_workbook(db_file_path)
        sheet = wb.worksheets[0]
        next_row_position = sheet.max_row + 1
    else:
        wb = workbook.Workbook()
        sheet = wb.worksheets[0]
        next_row_position = 1

    user = sheet.cell(next_row_position, 1)
    user.value = username

    pwd = sheet.cell(next_row_position, 2)
    pwd.value = md5(password)

    ctime = sheet.cell(next_row_position, 3)
    ctime.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save(db_file_path)


def run():
    while True:
        username = input("请输入用户名：")
        if username.upper() == "Q":
            break
        password = input("请输入密码：")
        register(username, password)


if __name__ == '__main__':
    run()